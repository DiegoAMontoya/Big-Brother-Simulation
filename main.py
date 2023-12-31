import random as ran
import time
import vetos as v
import hoh as h
vote_record = {}
evicted = []
HOHs = []
vetos = []
init_noms = []
fin_noms = []
elim_votes = []
tot_votes = []



contestants = ['Blake','Camila','Dav','Diego',
                'Dylan','Eliana','Faith','Gabby',
                'Ian','Jaden','Jaeden','Kaitlyn',
                'Kate','Madison','Miles','Ryan',
                'Richard','Sara','Sydney','Yogi']


numcon =9# ran.randint(9,18)
HOH = False
while True:
    ran.shuffle(contestants)
    rem_con = contestants[:numcon]
    if 'Diego' in rem_con:
        rem_con.sort()
        contestants.sort()
        print("{}\n".format(rem_con))
        break
for i in contestants:
    vote_record[i] = []
jury_num = (len(rem_con)-2)//2
if jury_num %2 != 1:
    jury_num += 1
    
week = 0
while len(rem_con)!=3:
    week +=1
    print("Week {}:\n".format(week))
    if len(rem_con)==jury_num + 2:
        input("We have reached the jury stage!\n")
    #HOH competition    
    HOH_elig = rem_con[:]
    if HOH in HOH_elig:
        HOH_elig.remove(HOH)
    HOH = h.hoh(HOH_elig)
    HOHs.append(HOH)
    input("This week's HOH is {}!\n".format(HOH))
    #initial nominees
    nom_elig = rem_con[:]
    nom_elig.remove(HOH)
    ran.shuffle(nom_elig)
    noms = nom_elig[0:2]
    noms.sort()
    init_noms.append(noms.copy())
    input("{} has nominated {} and {} for eviction.\n".format(HOH, noms[0],noms[1]))
    
    #veto selection and comp
    veto_play = [HOH]+noms
    elig_veto = rem_con[:]
    ran.shuffle(elig_veto)
    for i in elig_veto:
        if i not in veto_play:
            veto_play.append(i)
            if len(veto_play)==6:
                break
    input("Playing in the veto is {}.\n".format(veto_play))
    v_win = v.veto(veto_play)
    vetos.append(v_win)
    input("{} has won the power of veto!\n".format(v_win))
    
    #veto ceremony and rep nom
    print("{} has decided to...".format(v_win))
    time.sleep(1)
    if v_win != HOH:
        #1 = use, 0 = not using
        use = ran.randint(0, 1)
        if (use == 1 and len(rem_con) !=4) or v_win in noms:
            save = v_win
            if v_win in noms:
                noms.remove(v_win)
                input("use the power to save themself.\n")
            else:
                select = ran.randint(0,1)
                save = noms[select]
                noms.remove(save)
                input("use the power to save {}.\n".format(save))
            imm = noms + [save] + [v_win]
            ran.shuffle(nom_elig)
            for i in nom_elig:
                if i not in imm:
                    noms.append(i)
                    break
            input("{} has nominated {} in their place.\n".format(HOH,i))
        else:
            input("not use the power of veto.\n")
    else:
        input("not use the power of veto.\n")
    noms.sort()
    fin_noms.append(noms)  
    input("The final nominees are {} and {}.\n".format(noms[0],noms[1]))   
    
    #voting
    print("We will now begin voting.\n")
    time.sleep(1)
    vote_num = []
    votes = {}
    for i in noms:
        vote_num.append(0)
        votes[i] = []
    for i in rem_con:
        if i in noms:
            vote_record[i].append('NOM')
        elif i == HOH:
            vote_record[i].append('HOH')
        else:
            vote = ran.randint(0,len(noms)-1)
            vote_num[vote]+=1
            vote_record[i].append(noms[vote])
            votes[noms[vote]].append(i)
            
    #HOH vote if tied
    if vote_num[0]==vote_num[1]:
        vote = ran.randint(0,len(noms)-1)
        vote_num[vote]+=1
        vote_record[HOH][-1]= noms[vote]+'*'  
        votes[noms[vote]].append(HOH)
        input("There is a tie. {}, as HOH, will break the tie.\n".format(HOH))
        
    #removes most voted for nominee
    elimdex = vote_num.index(max(vote_num))
    evicted.append(noms[elimdex])
    rem_con.remove(noms[elimdex])
    vote_record[noms[elimdex]].append("Evicted\n(Week {})".format(week))
    vote_num.sort()
    input("By a vote of {} - {}...\n".format(vote_num[0],vote_num[1]))
    print("{} has been evicted from the big brother house.\n".format(noms[elimdex]))
    time.sleep(1)
    input("{}: {}\n{}: {}\n".format(noms[0],votes[noms[0]],noms[1],votes[noms[1]]))
    print("-----------\n")
    
    elim_votes.append(max(vote_num))
    tot_votes.append(sum(vote_num))    
          
   
#%%final 3
input("Welcome to the final 3!\n")
input("{} compete in part 1 of the final HOH.\n".format(rem_con))
ran.shuffle(rem_con)
p1_w =  rem_con[0]
input("{} wins part 1!\n".format(p1_w))
p2_players = rem_con[1:3]
ran.shuffle(p2_players)
p2_w = p2_players[0]
input("{} wins part 2!\n".format(p2_w))
p3_players = [p1_w,p2_w]
input("{} and {} compete in part 3.\n".format(p3_players[0],p3_players[1]))
ran.shuffle(p3_players)
HOH = p3_players[0]
HOHs.append(HOH)
input("{} is the final HOH!\n".format(HOH))
noms = rem_con[:]
noms.sort()
noms.remove(HOH)
init_noms.append(noms)

for i in noms:
    vote_record[i].append('NOM')
vote = ran.randint(0,len(noms)-1)
vote_num[vote]+=1
vote_record[HOH].append(noms[vote]+'*')
evicted.append(noms[vote])
rem_con.remove(noms[vote])
vote_record[noms[vote]][-1] = "Evicted\n(Week {})".format(week+1)

input("{} votes to evict...\n".format(HOH))
print("{}.\n".format(noms[vote]))
time.sleep(1)
input("{} is the last person to be evicted from the big brother house.\n".format(noms[vote]))
elim_votes.append(1)
tot_votes.append(1) 

#%%finale
finalist = rem_con[:]
jury = evicted[len(evicted)-jury_num:]
print("Please welcome back our jury!\n")
input("{}\n".format(jury))
for i in range(len(evicted)):
    if len(vote_record[rem_con[0]])-len(vote_record[evicted[i]]):
        for c in range(len(vote_record[rem_con[0]])-len(vote_record[evicted[i]])):
            vote_record[evicted[i]].append('    ')
print("The jury votes...\n")
time.sleep(1)            
vote_num = [0 for i in finalist]
for i in jury:
    vote = ran.randint(0,len(finalist)-1)
    vote_num[vote]+=1
    vote_record[i].append(finalist[vote])
            

elimdex = vote_num.index(min(vote_num))
for i in rem_con:
    if i != finalist[elimdex]:
        vote_record[i].append('Winner')
    else:
        vote_record[i].append('Runner-up')
evicted.append(finalist[elimdex])
rem_con.remove(finalist[elimdex])
  
vote_num.sort()
input("By a vote of {} - {}, the winner is...\n".format(vote_num[0],vote_num[1]))  
time.sleep(1)
print("{}!!!\n".format(rem_con[0]))
time.sleep(1)

#%%prep for spreadsheet
final_plac = rem_con
for i in range(len(evicted)):
    final_plac.append(evicted[-i-1])
for i in range(len(evicted)):
    if len(vote_record[final_plac[0]])-len(vote_record[evicted[i]]):
        for c in range(len(vote_record[final_plac[0]])-len(vote_record[evicted[i]])):
            vote_record[evicted[i]].append('    ')

records = [HOHs,init_noms,vetos,fin_noms]
for i in records:
    if len(vote_record[final_plac[0]])-len(i):
        for c in range(len(vote_record[final_plac[0]])-len(i)):
            i.append('    ')        

#make spreadsheet export
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
wb = Workbook()
sheet = wb.active
episodes = list(range(1,len(vote_record[rem_con[0]])+1))
alphabet = ['B','C','D','E','F','G','H','I','J','K','L','M','N',
            'O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AC']    
#colors for the spreadsheet
trowFill = PatternFill(start_color='00EAECF0',
                   end_color='00EAECF0',
                   fill_type='solid')
voteFill = PatternFill(start_color='00F8F9Fa',
                   end_color='00F8F9Fa',
                   fill_type='solid')
HOHFill = PatternFill(start_color='00ccffcc',
                   end_color='00ccffcc',
                   fill_type='solid')
nomFill = PatternFill(start_color='00959ffd',
                   end_color='00959ffd',
                   fill_type='solid')
naFill = PatternFill(start_color='00A9A9A9',
                   end_color='00A9A9A9',
                   fill_type='solid')
evFill = PatternFill(start_color='00fa8072',
                   end_color='00fa8072',
                   fill_type='solid')
winnerFill = PatternFill(start_color='0073fb76',
                   end_color='0073fb76',
                   fill_type='solid')
ruFill = PatternFill(start_color='00d1e8ef',
                   end_color='00d1e8ef',
                   fill_type='solid')

sheet["A2"] = "Head of\nHousehold"
sheet["A3"] = "Nominations\n(pre-veto)"
sheet["A4"] = "Veto\nWinner"
sheet["A5"] = "Nominations\n(post-veto)" 
#format the first column
for i in range(1,6):
    sheet["A{}".format(i)].fill = trowFill
    sheet["A{}".format(i)].font = Font(bold=True)
    sheet["A{}".format(i)].alignment = Alignment(horizontal="center")
rows = {}
columns = {}

for i in range(len(final_plac)):
    rows[final_plac[i]] = i+7
for i in range(len(alphabet)):
    columns[i] = alphabet[i]
#add in contents of first row and format
for i in range(len(episodes)):
    sheet["{}1".format(columns[i])] = "Week {}".format(episodes[i])    
    sheet["{}1".format(columns[i])].fill = trowFill
    sheet["{}1".format(columns[i])].font = Font(bold=True)
    sheet["{}1".format(columns[i])].alignment = Alignment(horizontal="center")
    for c in range(len(records)):
        if records[c]==init_noms or records[c]==fin_noms:
            sheet["{}{}".format(columns[i],c+2)] = "{}\n{}".format(records[c][i][0],records[c][i][1])
        else:
            sheet["{}{}".format(columns[i],c+2)] = str(records[c][i])
        sheet["{}{}".format(columns[i],c+2)].alignment = Alignment(horizontal="center")
        if str(records[c][i]) == '    ':
            sheet["{}{}".format(columns[i],c+2)].fill = naFill
        else:
            sheet["{}{}".format(columns[i],c+2)].fill = voteFill
sheet.merge_cells("{}1:{}1".format(columns[i-1],columns[i]))
#player additions            
for i in range(len(final_plac)):
    sheet["A{}".format(rows[final_plac[i]])] = final_plac[i]
    sheet["A{}".format(rows[final_plac[i]])].fill = trowFill
    sheet["A{}".format(rows[final_plac[i]])].font = Font(bold=True)
    sheet["A{}".format(rows[final_plac[i]])].alignment = Alignment(horizontal="center")  
     
    for c in range(len(vote_record[final_plac[i]])):
        sheet["{}{}".format(columns[c],rows[final_plac[i]])] = vote_record[final_plac[i]][c]
        sheet["{}{}".format(columns[c],rows[final_plac[i]])].alignment = Alignment(horizontal="center")
        if vote_record[final_plac[i]][c] == 'NOM':
            sheet["{}{}".format(columns[c],rows[final_plac[i]])] = 'Nominated'
            sheet["{}{}".format(columns[c],rows[final_plac[i]])].font = Font(italic=True)
            sheet["{}{}".format(columns[c],rows[final_plac[i]])].fill = nomFill
        elif vote_record[final_plac[i]][c] == 'HOH' or '*' in vote_record[final_plac[i]][c]:
            if '*' not in vote_record[final_plac[i]][c]:
                sheet["{}{}".format(columns[c],rows[final_plac[i]])] = 'Head of\nHousehold'
            sheet["{}{}".format(columns[c],rows[final_plac[i]])].font = Font(italic=True)
            sheet["{}{}".format(columns[c],rows[final_plac[i]])].fill = HOHFill
        elif vote_record[final_plac[i]][c] == 'Winner':
            sheet["{}{}".format(columns[c],rows[final_plac[i]])].fill = winnerFill  
            sheet["{}{}".format(columns[c],rows[final_plac[i]])].font = Font(bold=True)  
        elif vote_record[final_plac[i]][c] == 'Runner-up':
            sheet["{}{}".format(columns[c],rows[final_plac[i]])].fill = ruFill              
        elif vote_record[final_plac[i]][c] == '    ' or 'Evicted' in vote_record[final_plac[i]][c]:
            sheet["{}{}".format(columns[c],rows[final_plac[i]])].fill = evFill
        else:
            sheet["{}{}".format(columns[c],rows[final_plac[i]])].fill = voteFill
    
sheet["A{}".format(rows[final_plac[-1]]+2)] = "Evicted"
sheet["A{}".format(rows[final_plac[-1]]+2)].fill = trowFill
sheet["A{}".format(rows[final_plac[-1]]+2)].font = Font(bold=True)
sheet["A{}".format(rows[final_plac[-1]]+2)].alignment = Alignment(horizontal="center")
sheet.merge_cells("A{}:A{}".format(rows[final_plac[-1]]+2,rows[final_plac[-1]]+3))

for i in range(len(tot_votes)):
    sheet["{}{}".format(columns[i],rows[final_plac[-1]]+2)] = "{}\n{} of {} votes\nto evict".format(final_plac[-1-i],elim_votes[i],tot_votes[i])
    sheet["{}{}".format(columns[i],rows[final_plac[-1]]+2)].font = Font(bold=True)
    sheet["{}{}".format(columns[i],rows[final_plac[-1]]+2)].alignment = Alignment(horizontal="center")
    sheet["{}{}".format(columns[i],rows[final_plac[-1]]+2)].fill = evFill
    sheet.merge_cells("{}{}:{}{}".format(columns[i],rows[final_plac[-1]]+2,columns[i],rows[final_plac[-1]]+3))

sheet["{}{}".format(columns[i+1],rows[final_plac[-1]]+2)] = "{}\n{} votes\nto win".format(final_plac[1],vote_num[0])
sheet["{}{}".format(columns[i+1],rows[final_plac[-1]]+2)].font = Font(bold=True)
sheet["{}{}".format(columns[i+1],rows[final_plac[-1]]+2)].alignment = Alignment(horizontal="center")
sheet["{}{}".format(columns[i+1],rows[final_plac[-1]]+2)].fill = ruFill

sheet["{}{}".format(columns[i+1],rows[final_plac[-1]]+3)] = "{}\n{} votes\nto win".format(final_plac[0],vote_num[1])
sheet["{}{}".format(columns[i+1],rows[final_plac[-1]]+3)].font = Font(bold=True)
sheet["{}{}".format(columns[i+1],rows[final_plac[-1]]+3)].alignment = Alignment(horizontal="center")
sheet["{}{}".format(columns[i+1],rows[final_plac[-1]]+3)].fill = winnerFill

wb.save(filename="bbsim.xlsx")    